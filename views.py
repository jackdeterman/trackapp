import json
from datetime import datetime

from pprint import pprint

from django.core.paginator import Paginator
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth import decorators
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.db.models import Q
from django.http import JsonResponse
from django.shortcuts import (
    HttpResponse, HttpResponseRedirect, render, redirect, get_object_or_404)
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook

from .models import *
from .forms import *
from .event_dict import EVENT_DICT


@login_required
def index(request):

    return render(request, "index.html")

@login_required
def load_spreadsheet(request):

    if request.method == "POST":
        upload_form = UploadForm(request.POST, request.FILES)
        if not upload_form.is_valid():
            raise Exception("Error")

        team = upload_form.cleaned_data['team']

        season_name  = upload_form.cleaned_data['season']
        try:
            season = Season.objects.get(name=season_name)
        except Season.DoesNotExist:
            print(f"Creating season {season_name}")
            season = Season(
                name=season_name,
            )
            season.save()


        wb = load_workbook(upload_form.cleaned_data['file'])
        sheet = wb.active

        with transaction.atomic():
            headers = []
            for header in sheet[1]:
                headers.append(header.value.lower())

            total = 0
            users = set()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row = dict(zip(headers, row))

                first_name = row['first name'].strip()
                last_name = row['last name'].strip()
                username = f"{first_name.lower()}.{last_name.lower()}"

                try:
                    user = User.objects.get(username=username)
                except User.DoesNotExist:
                    print(f"Creating {username}")
                    user = User(
                        username=username,
                        first_name=first_name,
                        last_name=last_name
                    )
                    user.save()
                users.add(user)

                event_name = row['event'].strip()
                if event_name in EVENT_DICT:
                    if EVENT_DICT[event_name] != '':
                        event_name = EVENT_DICT[event_name]

                unit = 'inches'
                if 'meter' in event_name.lower():
                    unit = 'seconds'
                elif 'smr' in event_name.lower():
                    unit = 'seconds'
                elif 'hurdle' in event_name.lower():
                    unit = 'seconds'
                elif 'hurdle' in event_name.lower():
                    unit = 'seconds'
                elif 'yard' in event_name.lower():
                    unit = 'seconds'
                elif 'mile' in event_name.lower():
                    unit = 'seconds'
                elif 'medley' in event_name.lower():
                    unit = 'seconds'

                try:
                    event = Event.objects.get(name=event_name)
                except Event.DoesNotExist:
                    print(f"Creating {event_name}")
                    event = Event(
                        name=event_name,
                        unit=unit
                    )
                    event.save()

                if 'meet' in row:
                    meet_name = row['meet']
                    meet_date = row['date']
                elif 'opponent' in row:
                    meet_name = row['opponent']
                    meet_date = row['date']
                else:
                    meet_name = row['date']
                    date_str = meet_name.split(' ')[0]
                    date_str += '/2019'
                    meet_date = datetime.strptime(date_str, "%m/%d/%Y").date()

                try:
                    meet = Meet.objects.get(description=meet_name, date=meet_date)
                except Meet.DoesNotExist:
                    print(f"Creating {meet_name}")
                    meet = Meet(
                        description=meet_name,
                        date=meet_date,
                        team=team,
                        season=season
                    )
                    meet.save()

                performance = row['performance']
                if isinstance(performance, str):
                    if "-" in performance:
                        feet, inches = performance.split("-")
                        feet = float(feet)
                        inches = float(inches)
                        performance = (12.0 * feet) + inches
                    elif ':' in performance:
                        try:
                            pt = datetime.strptime(performance,'%M:%S.%f')
                        except:
                            print(f"Skipping {performance}")
                            continue
                        performance = pt.second + pt.minute*60.0 + pt.hour*3600.0 + (pt.microsecond/1000000.0)

                try:
                    performance = float(performance)
                except:
                    print(f"Skipping {performance}")
                    continue

                try:
                    result = Result.objects.get(
                        meet=meet,
                        event=event,
                        athlete=user,
                        result=performance
                    )
                except Result.DoesNotExist:
                    if 'fat/ht/na' in row:
                        method = row['fat/ht/na']
                    else:
                        method = row['fat / hand']
                    result = Result(
                        meet=meet,
                        event=event,
                        athlete=user,
                        result=performance,
                        method=method
                    )
                    result.save()
                    total += 1

            # Recalc
            for user in users:
                calculate_result_stats(user)

            #raise Exception("Data Test.")


        print(f"{total} rows processed") 
        return redirect('load_spreadsheet')       
    else:
        upload_form = UploadForm()

    return render(request, "load_spreadsheet.html", {'form': upload_form})

def login_view(request):
    if request.method == "POST":

        # Attempt to sign user in
        email = request.POST["email"]
        password = request.POST["password"]
        user = authenticate(request, username=email, password=password)

        # Check if authentication successful
        if user is not None:
            login(request, user)
            return HttpResponseRedirect('/')
        else:
            return render(request, "login.html", {
                "message": "Invalid email and/or password."
            })
    else:
        return render(request, "login.html")

@login_required
def logout_view(request):
    logout(request)
    return HttpResponseRedirect(reverse("index"))

def register(request):
    if request.method == "POST":
        email = request.POST["email"]
        first = request.POST["first"]
        last = request.POST["last"]

        # Ensure password matches confirmation
        password = request.POST["password"]
        confirmation = request.POST["confirmation"]
        if password != confirmation:
            return render(request, "register.html", {
                "message": "Passwords must match."
            })

        # Attempt to create new user
        try:
            user = User.objects.create_user(email, email, password)
            user.first_name = first
            user.last_name = last
            user.save()
        except IntegrityError as e:
            print(e)
            return render(request, "register.html", {
                "message": "Email address already taken."
            })
        login(request, user)
        return HttpResponseRedirect(reverse("index"))
    else:
        return render(request, "register.html")

@login_required
def user_list(request):

    users = User.objects.all().order_by("last_name")
    if "q" in request.GET:
        query = request.GET["q"]
        users = users.filter(
            Q(last_name__icontains=query) |
            Q(first_name__icontains=query) |
            Q(username__icontains=query) |
            Q(email=query)
        )
    paginator = Paginator(users, 50)
    page_number = request.GET.get('page')
    users = paginator.get_page(page_number)

    return render(request, "user_list.html", {"users":users})

@login_required
def profile(request, user_id):

    user = User.objects.get(id=user_id)
    results = Result.objects.filter(athlete=user).order_by('meet__date')
    goals = user.goals.all()

    results_by_event = {}
    for result in results:
        if result.event in results_by_event:
            results_by_event[result.event].append(result)
        else:
            results_by_event[result.event] = [result]

    return render(request, "profile.html", {
        'user': user,
        'results':results,
        'results_by_event':results_by_event,
        'goals': goals,
        })

@login_required
def meets(request):
    meets = Meet.objects.all().order_by("date")

    return render(request, "meets.html", {"meets":meets})

@login_required
def meet(request, meet_id):

    meet = Meet.objects.get(id=meet_id)
    results = Result.objects.filter(meet=meet).order_by('result')

    results_by_event = {}
    athletes = set()
    for result in results:
        athletes.add(result.athlete)
        if result.event in results_by_event:
            results_by_event[result.event].append(result)
        else:
            results_by_event[result.event] = [result]
    
    return render(request, "meet.html", {
        'meet': meet,
        'results':results,
        'athletes': athletes,
        'results_by_event': results_by_event
        })

@login_required
def events(request):
    events = Event.objects.all().order_by("name")

    # for event in events:
    #     if event.name in ['Discus', 'Shot Put', 'Discus Relay', 'High Jump',
    #         'High Jump Relay', 'Javelin', 'Javelin Relay', 'Long Jump', 'Long Jump Relay',
    #         'Pole Vault', 'Shot Put Relay', 'Triple Jump', 'Triple Jump Relay']:
    #         event.unit = 'Feet--Inches'
    #         event.save()

    return render(request, "events.html", {"events":events})

@login_required
def event(request, event_id):

    event = Event.objects.get(id=event_id)
    results = Result.objects.filter(event=event).order_by("result")

    return render(request, "event.html", {
        'event':event,
        'results':results
        })

@login_required
def merge_event(request, event_id):

    event = Event.objects.get(id=event_id)

    if request.method=="POST":
        form = MergeEventForm(request.POST)
        if form.is_valid():
            survivor = form.cleaned_data['event']
            event.results.all().update(event=survivor)
            event.delete()
            print(f"Merging {event.id} into {survivor.id}")
            return redirect('event', survivor.id)
    else:
        form = MergeEventForm()

    return render(request, "merge_event.html", {
        "form": form,
        "event":event,
    })


@login_required
def add_result(request, user_id):

    user = User.objects.get(id=user_id)
    results = Result.objects.filter(athlete=user)
    events = Event.objects.all()
    meets = Meet.objects.all()

    if request.method=="POST":
        form = ResultForm(request.POST)
        if form.is_valid():
            form.save(commit=False)
            form.instance.athlete=user
            form.instance.save()
    else:
        form = ResultForm()

    return render(request, "add_result.html", {
        "form": form,
        "user":user,
        "results":results,
        "events":events,
        "meets":meets
    })

@login_required
def edit_result(request, result_id):

    result = Result.objects.get(id=result_id)
    user = result.athlete
    results = Result.objects.filter(athlete=user)

    if request.method=="POST":
        form = ResultForm(request.POST, instance=result)
        if form.is_valid():
            form.save()
            calculate_result_stats(user)
            return redirect("profile", user.id)
    else:
        form = ResultForm(instance=result)

    return render(request, "edit_result.html", {
        "form": form,
        "user":user,
        "results":results,
        "events":events,
        "meets":meets
    })

@login_required
def delete_result(request, result_id):

    result = Result.objects.get(id=result_id)
    user = result.athlete

    if request.method=="POST":
        form = ResultForm(request.POST, instance=result)
        if form.is_valid():
            result.delete()
        return redirect("profile", user.id)
    else:
        form = ResultForm(instance=result)

    return render(request, "delete_result.html", {
        "form": form,
        "user":user,
    })

@login_required
def edit_profile(request, user_id):
    user = User.objects.get(id=user_id)

    if request.method=="POST":
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            return redirect("profile", user.id)
    else:
        form = UserForm(instance=user)

    return render(request, "edit_profile.html", {
        "form": form,
        "user":user,
    })

@login_required
def search(request):
    pass
#to do

@login_required
def merge_athlete(request, user_id):

    user = User.objects.get(id=user_id)

    if request.method=="POST":
        form = MergeAthleteForm(request.POST)
        if form.is_valid():
            survivor = form.cleaned_data['user']
            user.results.all().update(athlete=survivor)
            user.delete()
            print(f"Merging {user.id} into {survivor.id}")
            return redirect('profile', survivor.id)
    else:
        form = MergeAthleteForm()

    return render(request, "merge_athlete.html", {
        "form": form,
        "user":user,
    })

@login_required
def create_season_goal(request, user_id):


    user = User.objects.get(id=user_id)

    if request.method=="POST":
        form = SeasonGoalForm(request.POST)
        if form.is_valid():
            form.save(commit=False)
            form.instance.creator = request.user
            form.instance.user = user
            form.instance.save()
            return redirect("profile", user.id)
    else:
        form = SeasonGoalForm()
    
    return render(request, "create_season_goal.html", {
        "form": form,
        "user":user,
    })

@login_required
def remove_season_goal(request, goal_id):

    goal = Goal.objects.get(id=goal_id)
    user = goal.user

    if request.method=="POST":
        form = SeasonGoalForm(request.POST, instance=goal)
        if form.is_valid():
            goal.delete()
        return redirect("profile", user.id)
    else:
        form = SeasonGoalForm(instance=goal)

    return render(request, "remove_season_goal.html", {
        "goal":goal,
        "user":user,
        "form":form
    })

def merge_meet(request, meet_id):

    meet = Meet.objects.get(id=meet_id)

    if request.method=="POST":
        form = MergeMeetForm(request.POST)
        if form.is_valid():
            survivor = form.cleaned_data['meet']
            meet.results.all().update(meet=survivor)
            meet.delete()
            print(f"Merging {meet.description} into {survivor.description}")
            return redirect('meet', survivor.id)
    else:
        form = MergeMeetForm()

    return render(request, "merge_meet.html", {
        "form": form,
        "meet":meet,
    })
def teams(request):
    teams = Team.objects.all().order_by("name")

    return render(request, "teams.html", {"teams":teams})

@login_required
def team(request, team_id):
    team = get_object_or_404(Team, id=team_id)

    return render(request, "team.html", {"team":team})

@login_required
def edit_team(request, team_id=None):
    if team_id:
        team = get_object_or_404(Team, id=team_id)
    else:
        team = Team()

    if request.method=="POST":
        form = TeamForm(request.POST, instance=team)
        if form.is_valid():
            form.save()
            return redirect('team', team.id)
        else:
            print(form.errors)
    else:
        form = TeamForm(instance=team)

    return render(request, "edit_team.html", {
        "form": form,
    })

@login_required
def add_coach(request, team_id):

    team = Team.objects.get(id=team_id)

    if request.method=="POST":
        form = CoachForm(request.POST, instance=team)
        if form.is_valid():
            form.save()
            return redirect("team", team.id)
    else:
        form = CoachForm(instance=team)
    
    return render(request, "add_coach.html", {
        "form": form,
        "team":team,
    })

@login_required
def remove_coach(request, coach_id, team_id):

    coach = User.objects.get(id=coach_id)
    team = Team.objects.get(id=team_id)
    team.coaches.remove(coach)
    return redirect("team", team.id)

@login_required
def add_athlete_to_team(request, team_id):
    
    team = Team.objects.get(id=team_id)

    if request.method=="POST":
        form = AthleteTeamForm(request.POST, instance=team)
        if form.is_valid():
            form.save()
            return redirect("team", team.id)
    else:
        form = AthleteTeamForm(instance=team)
    
    return render(request, "add_athlete_to_team.html", {
        "form": form,
        "team":team,
    })

@login_required
def remove_athlete_from_team(request, athlete_id, team_id):

    athlete = User.objects.get(id=athlete_id)
    team = Team.objects.get(id=team_id)
    team.athletes.remove(athlete)
    return redirect("team", team.id)

@login_required
def debug_page(request):
    seasons = [
        'Indoor 2018-2019',
        'Indoor 2019-2020',
        'Indoor 2020-2021',
        'Indoor 2021-2022',
        'Outdoor 2019',
        'Outdoor 2020',
        'Outdoor 2021',
        'Outdoor 2022',
    ]
    for season_name in seasons:
        season = Season(name=season_name)
        season.save()


    return HttpResponse("Done")
